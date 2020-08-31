from openpyxl import Workbook
import datetime
import time
wb = Workbook()

ws = wb.active

ws['A1'] = 42
ws['B1'] = "你好"+"tset"

ws.append([1,2,3])

ws['A3'] = datetime.datetime.now()

ws['A4'] = time.strftime('%Y{y}%m{m}%d{d}').format(y='年',m='月',d='日')

wb.save("e:\\sample.xlsx")